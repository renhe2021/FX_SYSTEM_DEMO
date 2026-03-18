# -*- coding: utf-8 -*-
"""
excel_export.py - Excel 报表导出（带样式）
============================================
将分析结果导出为格式化的 Excel 文件。
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ---------- 样式 ----------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="微软雅黑", bold=True, size=11)
NORMAL_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
AMT_FORMAT = '#,##0'

# 金额列关键字
AMT_COL_KEYWORDS = ("损益金额", "其中WX", "FIT")

# 分组颜色
GROUP_FILLS = [
    PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
    PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
]

COL_WIDTHS = {
    "损益归属": 18, "产品类型": 14, "所属业务": 18, "商户号": 16,
    "损益金额": 18, "其中WX": 18, "FIT": 18, "商户类型": 14, "原币种": 10,
}


def _is_amt_col(col_name: str) -> bool:
    return any(kw in col_name for kw in AMT_COL_KEYWORDS)


def _get_col_width(col_name: str) -> int:
    if col_name in COL_WIDTHS:
        return COL_WIDTHS[col_name]
    if _is_amt_col(col_name):
        return 18
    return 15


def _write_sheet(ws, sheet_title: str, data: pd.DataFrame):
    """向一个 worksheet 写入数据并应用样式。"""
    num_cols = len(data.columns)
    num_rows = len(data)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=sheet_title)
    title_cell.font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头 (第2行)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # 数据行
    for row_idx in range(num_rows):
        excel_row = row_idx + 3
        first_val = str(data.iloc[row_idx, 0]) if num_cols > 0 else ""
        row_is_total = any(kw in first_val for kw in ["合计", "小计", "汇总"])

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            if row_is_total:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
            else:
                cell.font = NORMAL_FONT

            col_name = data.columns[col_idx - 1]
            if _is_amt_col(col_name):
                cell.number_format = AMT_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # 列宽
    for col_idx, col_name in enumerate(data.columns, 1):
        width = _get_col_width(col_name)
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

    # 冻结表头
    ws.freeze_panes = ws.cell(row=3, column=1)


def export_analysis_to_excel(
    analysis_results: dict,
    period_label: str,
    output_path: str | Path,
):
    """将单月分析结果导出为 Excel。"""
    sheet_map = {
        "明细汇总表": "detail",
        "分渠道汇总": "by_entity",
        "分业务汇总": "by_business",
        "分产品汇总": "by_product",
        "分币种汇总": "by_currency",
        "分商户类型": "by_merchant_type",
        "渠道×业务": "by_entity_business",
        "渠道×币种": "by_entity_currency",
        "分商户号汇总": "by_merchant",
    }

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        for sheet_name, key in sheet_map.items():
            if key in analysis_results:
                data = analysis_results[key]
                full_name = f"{period_label}-{sheet_name}"
                if len(full_name) > 31:
                    full_name = full_name[:31]
                data.to_excel(writer, sheet_name=full_name, index=False, startrow=1)
                ws = writer.sheets[full_name]
                _write_sheet(ws, full_name, data)

    return str(output_path)


def export_multi_period_excel(
    period_results: list[tuple[str, dict]],
    combined_results: dict | None,
    comparison_dfs: dict | None,
    output_path: str | Path,
):
    """将多月份 + 累计 + 对比结果导出到同一个 Excel。"""
    sheet_map = {
        "明细汇总表": "detail",
        "分渠道汇总": "by_entity",
        "分业务汇总": "by_business",
        "分产品汇总": "by_product",
        "分币种汇总": "by_currency",
        "分商户号汇总": "by_merchant",
    }

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        # 月度对比 sheet
        if comparison_dfs:
            for comp_name, comp_df in comparison_dfs.items():
                sheet_name = f"月度对比-{comp_name}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                comp_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                ws = writer.sheets[sheet_name]
                _write_sheet(ws, sheet_name, comp_df)

        # 各月份
        for period_label, results in period_results:
            for sheet_name, key in sheet_map.items():
                if key in results:
                    data = results[key]
                    full_name = f"{period_label}-{sheet_name}"
                    if len(full_name) > 31:
                        full_name = full_name[:31]
                    data.to_excel(writer, sheet_name=full_name, index=False, startrow=1)
                    ws = writer.sheets[full_name]
                    _write_sheet(ws, full_name, data)

        # 累计
        if combined_results:
            for sheet_name, key in sheet_map.items():
                if key in combined_results:
                    data = combined_results[key]
                    full_name = f"累计-{sheet_name}"
                    if len(full_name) > 31:
                        full_name = full_name[:31]
                    data.to_excel(writer, sheet_name=full_name, index=False, startrow=1)
                    ws = writer.sheets[full_name]
                    _write_sheet(ws, full_name, data)

    return str(output_path)
